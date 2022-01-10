import requests
import time
import csv
import json
import re

# https://www.cnpython.com/qa/481686
# https://zhuanlan.zhihu.com/p/81009691

'''
主要介绍pandas中
（1）isin函数
（2）contains函数
（3）apply函数
（4）字符串中的in函数
（5）正则匹配

'''
import pandas as pd



df=pd.read_excel('./对比.xlsx')
print(type(df))
df['test A']=df['test A'].str.replace('、','')
df['test B']=df['test B'].str.replace('、','')
df=df.rename(columns={'test A':'A','test B':'B'},inplace=True)

data=pd.read_excel('./对比.xlsx')
for i in range(len(data)):
    a = set(data['test A'][i])
    b = set(data['test B'][i])
    c = a-b
    if c:
        r = '否'
    else:
        r = '是'
    data['是否覆盖'][i]=r



df.apply(lambda x: re.search(r'\b' + x.A + r'\b', x.B) is not None, axis=1)
df.apply(lambda x: x.A in x.B, axis=1)
df.apply(lambda row: row['a'] in row['b'], axis=1)




# coding=utf-8
from openpyxl import load_workbook

# 分割符号
SPLIT = "、"
# 输入文件路径
INPUT_FILE = "对比.xlsx"

workbook = load_workbook(INPUT_FILE)

sheet = workbook.active


def check_left_contain(str1, str2):
    a1 = str1.split(SPLIT)
    a2 = str2.split(SPLIT)
    s1 = set(a1)
    s2 = set(a2)
    s = s1 & s2
    return len(s1) == len(s)


i = 0
for row in sheet.rows:
    i = i + 1
    if i == 1:
        continue

    if row[0] is not None and row[1].value is not None:
        is_contain = check_left_contain(str(row[0].value), str(row[1].value))
        if is_contain is True:
            sheet.cell(row=i, column=3).value = '是'
        else:
            sheet.cell(row=i, column=3).value = '否'

workbook.save('结果.xlsx')
workbook.close()


